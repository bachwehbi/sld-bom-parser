import io
import json
import logging
import uuid
from datetime import datetime

import httpx
from fastapi import APIRouter, HTTPException
from fastapi.responses import StreamingResponse
from pydantic import BaseModel

from server.config import (
    get_config, get_workspace_client, exec_sql,
    MATCHES_TABLE, EXPORTS_TABLE, EXPORTS_PATH, MATCHING_JOB_ID,
    VOLUME_PATH, TABLE_NAME,
)

logger = logging.getLogger(__name__)
router = APIRouter()


# ── Request models ────────────────────────────────────────────────────────────

class MatchRequest(BaseModel):
    file_name:      str
    top_n:          int   = 3
    preferred_tier: str   = ""


class OverrideItem(BaseModel):
    component_idx:  int
    selected_reference: str
    status:         str = "OVERRIDDEN"   # ACCEPTED | OVERRIDDEN | SKIPPED


class OverrideRequest(BaseModel):
    file_name:  str
    overrides:  list[OverrideItem]


# ── Trigger matching ──────────────────────────────────────────────────────────

@router.post("/api/match")
def trigger_matching(req: MatchRequest):
    """Submit the reference matching job for a diagram."""
    w    = get_workspace_client()
    host = w.config.host.rstrip("/")
    hdrs = {**w.config.authenticate(), "Content-Type": "application/json"}

    payload = {
        "job_id": MATCHING_JOB_ID,
        "notebook_params": {
            "file_name":      req.file_name,
            "top_n":          str(req.top_n),
            "preferred_tier": req.preferred_tier,
        },
    }
    try:
        resp = httpx.post(f"{host}/api/2.1/jobs/run-now", headers=hdrs, json=payload, timeout=15)
        resp.raise_for_status()
        run_id = resp.json()["run_id"]
    except httpx.HTTPStatusError as e:
        logger.exception("Matching job trigger failed")
        raise HTTPException(500, f"Job trigger failed: {e.response.text[:200]}")

    return {"run_id": run_id, "file_name": req.file_name, "status": "submitted"}


# ── Check job run status ─────────────────────────────────────────────────────

@router.get("/api/match/run-status/{run_id}")
def get_run_status(run_id: int):
    """Poll a matching job run state (PENDING/RUNNING/SUCCEEDED/FAILED)."""
    w    = get_workspace_client()
    host = w.config.host.rstrip("/")
    hdrs = {**w.config.authenticate(), "Content-Type": "application/json"}
    try:
        resp = httpx.get(f"{host}/api/2.1/jobs/runs/get?run_id={run_id}", headers=hdrs, timeout=15)
        resp.raise_for_status()
        data  = resp.json()
        state = data.get("state", {})
        return {
            "run_id":        run_id,
            "life_cycle":    state.get("life_cycle_state"),
            "result_state":  state.get("result_state"),
            "state_message": state.get("state_message", ""),
        }
    except httpx.HTTPStatusError as e:
        raise HTTPException(500, f"Jobs API error: {e.response.text[:200]}")


# ── Get match results ─────────────────────────────────────────────────────────

@router.get("/api/matches/{file_name}")
def get_matches(file_name: str):
    """Return current reference_matches rows for a diagram."""
    safe = file_name.replace("'", "")
    try:
        rows = exec_sql(f"""
            SELECT component_idx, component_summary, suggested_references,
                   selected_reference, user_overridden, status, updated_at
            FROM {MATCHES_TABLE}
            WHERE file_name = '{safe}'
            ORDER BY component_idx
        """, max_rows=500)
    except Exception as e:
        logger.exception("get_matches failed")
        raise HTTPException(500, str(e))

    if not rows:
        return {"file_name": file_name, "status": "NOT_MATCHED", "matches": []}

    result = []
    for r in rows:
        refs_raw = r.get("suggested_references") or "[]"
        try:
            refs = json.loads(refs_raw) if isinstance(refs_raw, str) else refs_raw
        except Exception:
            refs = []
        result.append({
            "component_idx":       r.get("component_idx"),
            "component_summary":   r.get("component_summary"),
            "suggested_references": refs,
            "selected_reference":  r.get("selected_reference"),
            "user_overridden":     r.get("user_overridden"),
            "status":              r.get("status"),
            "updated_at":          r.get("updated_at"),
        })

    return {"file_name": file_name, "status": "MATCHED", "matches": result}


# ── Save user overrides ───────────────────────────────────────────────────────

@router.patch("/api/matches")
def save_overrides(req: OverrideRequest):
    """Persist user-selected references (batch update via individual MERGEs)."""
    safe_file = req.file_name.replace("'", "")
    now       = datetime.utcnow().isoformat()
    updated   = 0
    errors    = []

    for item in req.overrides:
        status = item.status if item.status in ("ACCEPTED", "OVERRIDDEN", "SKIPPED") else "OVERRIDDEN"
        overridden = "true" if status == "OVERRIDDEN" else "false"
        safe_ref   = (item.selected_reference or "").replace("'", "")
        try:
            exec_sql(f"""
                MERGE INTO {MATCHES_TABLE} AS t
                USING (SELECT {item.component_idx} AS component_idx) AS s
                ON t.file_name = '{safe_file}' AND t.component_idx = s.component_idx
                WHEN MATCHED THEN UPDATE SET
                    selected_reference = '{safe_ref}',
                    user_overridden    = {overridden},
                    status             = '{status}',
                    updated_at         = '{now}'
            """)
            updated += 1
        except Exception as e:
            errors.append({"component_idx": item.component_idx, "error": str(e)})

    return {"updated": updated, "errors": errors}


# ── Generate and download Excel ───────────────────────────────────────────────

@router.post("/api/export/{file_name}")
def generate_export(file_name: str):
    """Generate an Excel file, save to UC volume, record in exports table."""
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(500, "openpyxl not installed")

    safe = file_name.replace("'", "")

    # Load BOM components
    bom_rows = exec_sql(
        f"SELECT bom_json FROM {TABLE_NAME} WHERE file_name = '{safe}' AND status = 'SUCCESS'",
        max_rows=1,
    )
    if not bom_rows:
        raise HTTPException(404, f"No successful extraction for {file_name}")
    try:
        components = json.loads(bom_rows[0]["bom_json"] or "[]")
    except Exception:
        components = []

    # Load matches
    match_rows = exec_sql(f"""
        SELECT component_idx, component_summary, selected_reference, suggested_references,
               user_overridden, status
        FROM {MATCHES_TABLE}
        WHERE file_name = '{safe}'
        ORDER BY component_idx
    """, max_rows=500)

    # Build lookup: idx → match row
    match_by_idx = {}
    for m in match_rows:
        try:
            refs = json.loads(m.get("suggested_references") or "[]")
        except Exception:
            refs = []
        match_by_idx[int(m["component_idx"])] = {**m, "refs_list": refs}

    # ── Build Excel ───────────────────────────────────────────────────────────
    wb = openpyxl.Workbook()

    # Styles
    hdr_fill    = PatternFill("solid", fgColor="1F3864")
    hdr_font    = Font(color="FFFFFF", bold=True, size=10)
    hdr_align   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    green_fill  = PatternFill("solid", fgColor="C6EFCE")
    yellow_fill = PatternFill("solid", fgColor="FFEB9C")
    red_fill    = PatternFill("solid", fgColor="FFC7CE")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )

    def set_header(ws, headers, col_widths):
        for col, (h, w) in enumerate(zip(headers, col_widths), start=1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill    = hdr_fill
            cell.font    = hdr_font
            cell.alignment = hdr_align
            cell.border  = thin_border
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.row_dimensions[1].height = 30

    # ── Sheet 1: BOM + selected references ───────────────────────────────────
    ws1 = wb.active
    ws1.title = "BOM with References"
    ws1.freeze_panes = "A2"

    bom_headers = [
        "Tipo", "Calibre (A)", "Curva", "Poder Corte (kA)", "Polos",
        "Tensión (V)", "Sensibilidad (mA)", "Tipo Dif.", "Función",
        "Cuadro", "Circuito",
        "Referencia Seleccionada", "Descripción", "Gama", "Tier",
        "Precio Lista (€)", "Stock", "DC Stock", "ETA Reposición",
        "Modificado por usuario",
    ]
    bom_widths = [
        24, 10, 8, 14, 7,
        10, 14, 10, 12,
        14, 22,
        22, 36, 14, 10,
        14, 16, 14, 16,
        10,
    ]
    set_header(ws1, bom_headers, bom_widths)

    total_value = 0.0
    referenced_count = 0
    overridden_count = 0

    for idx, comp in enumerate(components):
        m        = match_by_idx.get(idx, {})
        sel_ref  = m.get("selected_reference") or ""
        refs_list = m.get("refs_list", [])

        # Find selected ref detail
        ref_detail = next((r for r in refs_list if r.get("reference") == sel_ref), refs_list[0] if refs_list else {})

        stock_status = ref_detail.get("stock_status", "")
        price        = ref_detail.get("list_price_eur")
        if price:
            total_value += float(price)

        row_data = [
            comp.get("Que és", ""),
            comp.get("Calibre (A)", ""),
            comp.get("Curva", ""),
            comp.get("Poder de Corte (kA)", ""),
            comp.get("Polos", ""),
            comp.get("Tensión (V)", ""),
            comp.get("Sensibilidad (mA)", ""),
            comp.get("Tipo (Diferencial)", ""),
            comp.get("Función (Reloj)", ""),
            comp.get("Cuadro", ""),
            comp.get("Circuito", ""),
            sel_ref,
            ref_detail.get("product_description", ""),
            ref_detail.get("range", ""),
            ref_detail.get("tier", ""),
            price,
            stock_status,
            ref_detail.get("distribution_center", ""),
            ref_detail.get("expected_date", ""),
            "Sí" if m.get("user_overridden") else "",
        ]

        row_num = idx + 2
        for col, val in enumerate(row_data, start=1):
            cell = ws1.cell(row=row_num, column=col, value=val)
            cell.border = thin_border
            cell.font   = Font(size=9)

        # Stock color coding
        stock_cell = ws1.cell(row=row_num, column=17)
        if stock_status == "IN_STOCK":
            stock_cell.fill = green_fill
        elif stock_status == "LOW_STOCK":
            stock_cell.fill = yellow_fill
        elif stock_status == "OUT_OF_STOCK":
            stock_cell.fill = red_fill

        if sel_ref:
            referenced_count += 1
        if m.get("user_overridden"):
            overridden_count += 1

    # ── Sheet 2: All alternatives ─────────────────────────────────────────────
    ws2 = wb.create_sheet("Alternatives")
    ws2.freeze_panes = "A2"

    alt_headers = [
        "Componente", "Cuadro", "Circuito",
        "Referencia", "Descripción", "Gama", "Tier", "Estado",
        "Precio Lista (€)", "Puntuación", "Stock", "DC Stock", "ETA",
    ]
    alt_widths = [26, 14, 22, 22, 36, 14, 10, 12, 14, 10, 16, 14, 16]
    set_header(ws2, alt_headers, alt_widths)

    alt_row = 2
    for idx, comp in enumerate(components):
        m         = match_by_idx.get(idx, {})
        refs_list = m.get("refs_list", [])
        tipo      = comp.get("Que és", "")
        cuadro    = comp.get("Cuadro", "")
        circuit   = comp.get("Circuito", "")
        for ref in refs_list:
            row_data = [
                tipo, cuadro, circuit,
                ref.get("reference", ""),
                ref.get("product_description", ""),
                ref.get("range", ""),
                ref.get("tier", ""),
                ref.get("status", ""),
                ref.get("list_price_eur"),
                ref.get("score"),
                ref.get("stock_status", ""),
                ref.get("distribution_center", ""),
                ref.get("expected_date", ""),
            ]
            for col, val in enumerate(row_data, start=1):
                cell = ws2.cell(row=alt_row, column=col, value=val)
                cell.border = thin_border
                cell.font   = Font(size=9)
            stock_cell = ws2.cell(row=alt_row, column=11)
            st = ref.get("stock_status", "")
            if st == "IN_STOCK":
                stock_cell.fill = green_fill
            elif st == "LOW_STOCK":
                stock_cell.fill = yellow_fill
            elif st == "OUT_OF_STOCK":
                stock_cell.fill = red_fill
            alt_row += 1

    # ── Save to bytes ─────────────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    xlsx_bytes = buf.read()

    # ── Upload to UC volume ───────────────────────────────────────────────────
    stem      = file_name.replace(".pdf", "").replace(".PDF", "")
    timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    export_id = str(uuid.uuid4())
    vol_path  = f"{EXPORTS_PATH}/{stem}_{timestamp}.xlsx"

    w    = get_workspace_client()
    host = w.config.host.rstrip("/")
    hdrs = w.config.authenticate()
    url  = f"{host}/api/2.0/fs/files{vol_path}?overwrite=true"
    try:
        upload_resp = httpx.put(
            url,
            headers={**hdrs, "Content-Type": "application/octet-stream"},
            content=xlsx_bytes,
            timeout=60,
        )
        upload_resp.raise_for_status()
    except httpx.HTTPStatusError as e:
        logger.warning(f"Could not save export to volume: {e}")
        # Continue — still return the file to the user

    # ── Record in exports table ───────────────────────────────────────────────
    try:
        user = "unknown"
        try:
            user = w.current_user.me().user_name or "unknown"
        except Exception:
            pass
        now_ts = datetime.utcnow().isoformat()
        exec_sql(f"""
            INSERT INTO {EXPORTS_TABLE} (
                export_id, file_name, exported_by, exported_at, volume_path,
                component_count, referenced_count, overridden_count, total_value_eur
            ) VALUES (
                '{export_id}', '{safe}', '{user}', '{now_ts}', '{vol_path}',
                {len(components)}, {referenced_count}, {overridden_count},
                {round(total_value, 2)}
            )
        """)
    except Exception as e:
        logger.warning(f"Could not record export: {e}")

    # ── Stream back to user ───────────────────────────────────────────────────
    fname = f"{stem}_bom_references_{timestamp}.xlsx"
    return StreamingResponse(
        io.BytesIO(xlsx_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


# ── Past exports ──────────────────────────────────────────────────────────────

@router.get("/api/exports/{file_name}")
def get_exports(file_name: str):
    """List past Excel exports for a diagram."""
    safe = file_name.replace("'", "")
    try:
        rows = exec_sql(f"""
            SELECT export_id, exported_by, exported_at, volume_path,
                   component_count, referenced_count, overridden_count, total_value_eur
            FROM {EXPORTS_TABLE}
            WHERE file_name = '{safe}'
            ORDER BY exported_at DESC
        """, max_rows=20)
    except Exception as e:
        logger.exception("get_exports failed")
        raise HTTPException(500, str(e))
    return {"file_name": file_name, "exports": rows}


@router.get("/api/exports/download/{export_id}")
def download_export(export_id: str):
    """Re-serve a previously generated Excel from the UC volume."""
    safe = export_id.replace("'", "")
    rows = exec_sql(
        f"SELECT volume_path, file_name FROM {EXPORTS_TABLE} WHERE export_id = '{safe}'",
        max_rows=1,
    )
    if not rows:
        raise HTTPException(404, "Export not found")

    vol_path  = rows[0]["volume_path"]
    file_name = rows[0]["file_name"]
    w         = get_workspace_client()
    host      = w.config.host.rstrip("/")
    hdrs      = w.config.authenticate()
    url       = f"{host}/api/2.0/fs/files{vol_path}"

    fname = vol_path.split("/")[-1]

    def stream():
        with httpx.stream("GET", url, headers=hdrs, timeout=30) as resp:
            if resp.status_code == 404:
                raise HTTPException(404, "Export file not found in volume")
            resp.raise_for_status()
            for chunk in resp.iter_bytes(chunk_size=65536):
                yield chunk

    return StreamingResponse(
        stream(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )
